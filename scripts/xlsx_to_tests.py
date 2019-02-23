#!/usr/bin/env python3
import unittest
import os
import subprocess
import openpyxl as op
import json
import argparse
import re
from pathlib import Path
from dataclasses import dataclass
from typing import Callable, Union, List, Any, Optional, Tuple
from inspect import signature
from collections import defaultdict
from itertools import product
from shutil import which

_debug = False


def debug(format_string, *args):
    if _debug:
        print(format_string.format(*args))


_test_file_location_template = "packages/{package}/src/__tests__/{type_plural}/{test_file}"
# TODO not implemented
_mock_file_location_template = "packages/{package}/src/__mocks__/{mock_file}"
_file_suffix = ".spec.ts"
_expected_packages = [
    "core",
    "fs",
    "imports",
    "imports-fs",
]

_data_regex = r"const data = \[[\s\S]*?\];.*"
_data_template = '''const data = [
{test_data}
]; // autogenerated'''


_lambda_tag = "Lambdas"
PartTypes = ["Parser", "Resolver"]
RawInput_t = Union[Optional[str], List[Optional[str]]]
Input_t = Any
Output_t = Optional[str]
Transform_t = Callable[[Input_t], Input_t]


@dataclass
class SheetEntry:
    "Class representing an entry in any of the sheets, with some extra functionality"
    PartType: str
    PartName: str
    Input: Input_t
    Output: Output_t
    Transform: dict

    def transform(self, name: str, transform: Callable[[Input_t], Input_t]):
        if name not in self.Transform.values():
            return

        for target, v in self.Transform.items():
            if v == name:
                transmorphed = transform(self.__getattribute__(target))
                self.__setattr__(target, transmorphed)

    def predict_location(self, package_name: str):
        return _test_file_location_template.format(**{
            "package": package_name,
            "type_plural": self.PartType.lower() + "s",
            "test_file": self.PartName + self.PartType + _file_suffix,
        })


@dataclass
class RawSheetEntry:
    "Raw row, consisting only of string fields"
    PartType: str
    PartName: str
    Input: str
    Output: str
    Transform: str

    def parse(self) -> SheetEntry:
        pt = self.PartType
        pn = self.PartName
        inp = json.loads(self.Input)
        outp = json.loads(self.Output)
        trans = json.loads(
            self.Transform) if self.Transform is not None else {}

        return SheetEntry(pt, pn, inp, outp, trans)


def generate_data_from_sheet_entries(data: List[SheetEntry]) -> str:
    indentation = "    "
    data_lines = []
    for _datum in data:
        datum = _datum  # type: SheetEntry
        line_obj = [datum.Input, datum.Output]
        line_str = indentation + json.dumps(line_obj)
        data_lines.append(line_str)

    joined = ",\n".join(data_lines)  # with indentation
    return _data_template.format(test_data=joined)


def write_new_test_file(filepath: str, data: List[SheetEntry]):
    parent_dir = Path(Path(filepath).parent)
    if not parent_dir.exists():
        parent_dir.mkdir(parents=True)

    with open(filepath, "w") as fp:
        fp.write(generate_data_from_sheet_entries(data))


def write_existing_test_file(filepath: str, data: List[SheetEntry]):
    fr = open(filepath, "r")
    file_contents = fr.read()
    fr.close()
    pat_o = re.compile(_data_regex)

    new_contents = file_contents
    if pat_o.match(file_contents):
        new_contents = pat_o.sub(
            generate_data_from_sheet_entries(data), file_contents)
    else:
        # if pattern failed, then the new data will be inserted at the top of the file
        # we are prettierizing the code anyways
        new_contents = "{}\n{}".format(
            generate_data_from_sheet_entries(data), file_contents)

    fw = open(filepath, "w")
    fw.write(new_contents)


def parse_xlsx_sheet(sheet: op.worksheet.Worksheet) -> Tuple[List[RawSheetEntry], dict]:
    take_at_most_n_arguments = len(signature(
        RawSheetEntry.__init__).parameters.items()) - 1  # minus 'self'
    rows = []
    lambs = {}
    last_row = 0
    for row in sheet:
        last_row += 1
        if row[0].value is None:
            continue
        if row[0].value == _lambda_tag:
            break
        raw_row = RawSheetEntry(
            *[x.value for x in row[:take_at_most_n_arguments]])
        rows.append(raw_row)

    # if there was a Lambdas declaration then we have broken the loop
    # else we've iterated through all the empty rows so this iterator returns nothing
    for row in sheet.iter_rows(min_row=last_row + 1):
        lambs[row[0].value] = row[1].value

    return rows[1:], dict([(k, v) for k, v in lambs.items() if k is not None and v is not None])


def traverse_file_tree(cwd=".", file_suffix=_file_suffix):
    test_files = []
    for root, _, files in os.walk(cwd):
        for f in files:
            sufs = Path(f).suffixes
            try:
                ext = sufs[-1]
                inf = sufs[-2]
                if inf + ext == file_suffix:  # I can do better than that
                    path_to_f = (Path(root) / Path(f)).as_posix()
                    test_files.append(path_to_f)
            except IndexError:
                continue
    return test_files


_prettier_confs = [
    "\.prettierrc\.(yaml|yml|json|toml)", "\.?prettier(\.config|rc)\.js"]


def prettier(files_to_check=None):
    if files_to_check is not None:
        debug("Prettierizing {}", files_to_check)

    checkers = [re.compile(pattern) for pattern in _prettier_confs]
    here = "."
    onlyfiles = [f for f in os.listdir(here) if os.path.isfile(f)]

    config_file = ""
    for f, ch in product(onlyfiles, checkers):
        if ch.match(f):
            # found match
            config_file = f
            break
    else:
        # no prettier config detected
        return

    # we want to use yarn, because 'npm run' sucks
    # for this we check if 'yarn' exists in path
    if which("yarn") is None:
        return

    command = ["yarn", "prettier", "--config", config_file, "--write"]
    if files_to_check is not None:
        command.extend(files_to_check)

    out = subprocess.run(command)
    debug("\"{}\" returned {}", " ".join(command), out.returncode)


def main(input_file, output_dir=None, only_data=False, sheets=None, **kwargs):
    # this flags gets the evaluation of the cells (but does not evaluate them! those values are cached somewhere in .xlsx)
    wb = op.load_workbook(input_file, data_only=True)
    existing_test_files = set(traverse_file_tree())
    all_test_files = set(existing_test_files)

    for sheetname in wb.sheetnames:
        if sheets is not None and sheetname.lower() not in sheets:
            continue

        if sheetname not in _expected_packages:
            continue

        raw_rows, lambdas = parse_xlsx_sheet(wb[sheetname])
        value_rows = [row.parse() for row in raw_rows]
        # evaluating both, because keys are strings describing a string, e.g. '"cwd"'
        evaluated_lambdas = dict([(eval(k), eval(v))
                                  for k, v in lambdas.items()])
        for k, v in evaluated_lambdas.items():
            # fuck computer resources; electricity is cheap anyways
            for row in value_rows:
                row.transform(k, v)

        files_to_data = defaultdict(list)
        planned_test_files = set()
        for row in value_rows:
            file_path = row.predict_location(sheetname)
            planned_test_files.add(file_path)
            files_to_data[file_path].append(row)

        missing_test_files = planned_test_files.difference(existing_test_files)
        rest_of_test_files = planned_test_files.difference(missing_test_files)

        all_test_files = all_test_files.union(planned_test_files)

        for file_p in missing_test_files:
            write_new_test_file(file_p, files_to_data[file_p])

        for file_p in rest_of_test_files:
            write_existing_test_file(file_p, files_to_data[file_p])

    # do some extra magic
    # if prettier exists, run it at the root of the project
    prettier(all_test_files)


_package_json = "package.json"


def always_exec_relative_to_package_json():
    abspath = os.path.abspath(__file__)
    dir_name = os.path.dirname(abspath)

    cwd_p = Path(dir_name)
    found = None
    while not found and cwd_p.as_posix() != cwd_p.root:
        for f in cwd_p.iterdir():
            if Path(f).parts[-1] == _package_json:
                found = f
                break
        else:
            cwd_p = cwd_p.parent

    project_root_dir = Path(found).parent
    debug("Found {}. Changing dir to {}", found, project_root_dir)
    os.chdir(project_root_dir)


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("file", help="input M$ Excel file", type=str)
    parser.add_argument(
        "--sheets", help="choose sheets only from given comma-delimited list (standard, unix one)", type=str)
    parser.add_argument(
        "--debug", help="print extra information", action="store_true")

    args = parser.parse_args()
    func_args = {}
    if getattr(args, "debug"):
        _debug = True
    func_args["input_file"] = getattr(args, "file")
    if getattr(args, "sheets"):
        csl: str = getattr(args, "sheets")
        func_args["sheets"] = [x.lower() for x in csl.split(",")]

    debug("func_args: {}", func_args)

    always_exec_relative_to_package_json()
    main(**func_args)


#
#
#
#
# tests


class ParsingRawRowsTestCase(unittest.TestCase):
    def setUp(self):
        self.test_cases = [
            (RawSheetEntry(
                "Resolver", "Fs", r'[null, ".", {"property": "value"}]', "null", r'{"Input": "cwd"}'
            ), SheetEntry(
                "Resolver", "Fs", [None, ".", {
                    "property": "value"}], None, {"Input": "cwd"}
            )),
            (RawSheetEntry(
                "Resolver", "Fs", "null", "null", None
            ), SheetEntry(
                "Resolver", "Fs", None, None, {}
            )),
            (RawSheetEntry(
                "Resolver", "Fs", r"{}", r"{}", None
            ), SheetEntry(
                "Resolver", "Fs", {}, {}, {}
            ))
        ]

    def assert_objects_match_after_parse(self, obj1: RawSheetEntry, obj2: SheetEntry):
        parsed = obj1.parse()
        self.assertEqual(parsed.PartType, obj2.PartType)
        self.assertEqual(parsed.PartName, obj2.PartName)

        if obj2.Input.__class__ == list:
            self.assertListEqual(parsed.Input, obj2.Input)
        elif obj2.Input.__class__ == dict:
            self.assertDictEqual(parsed.Input, obj2.Input)
        else:
            self.assertEqual(parsed.Input, obj2.Input)

        if obj2.Output.__class__ == list:
            self.assertListEqual(parsed.Output, obj2.Output)
        elif obj2.Output.__class__ == dict:
            self.assertDictEqual(parsed.Output, obj2.Output)
        else:
            self.assertEqual(parsed.Output, obj2.Output)

        self.assertDictEqual(parsed.Transform, obj2.Transform)

    def test_pairs(self):
        for i, pair in enumerate(self.test_cases):
            with self.subTest(i=i):
                self.assert_objects_match_after_parse(pair[0], pair[1])


class EvaluatingRowsTestCase(unittest.TestCase):
    def setUp(self):
        self.raw_row = RawSheetEntry(
            "Resolver", "Fs", r'["string", "."]', "null", r'{"Input": "cwd"}'
        )

        self.transform = ("cwd", lambda x: [x[0], {"cwd": x[1]}])

    def test_parsing(self):
        parsed_row = self.raw_row.parse()

        parsed_row.transform(*self.transform)

        self.assertEqual(parsed_row.PartType, "Resolver")
        self.assertEqual(parsed_row.PartName, "Fs")
        self.assertListEqual(parsed_row.Input, ["string", {"cwd": "."}])
        self.assertEqual(parsed_row.Output, None)
        self.assertDictEqual(parsed_row.Transform, {"Input": "cwd"})


class EvaluatingRowsMismatchTestCase(unittest.TestCase):
    def setUp(self):
        self.raw_row = RawSheetEntry(
            "Resolver", "Fs", r'["string", "."]', "null", r'{"Input": "mistake"}'
        )

        self.transform = ("cwd", lambda x: [x[0], {"cwd": x[1]}])

    def test_parsing(self):
        parsed_row = self.raw_row.parse()

        parsed_row.transform(*self.transform)

        self.assertEqual(parsed_row.PartType, "Resolver")
        self.assertEqual(parsed_row.PartName, "Fs")
        self.assertListEqual(parsed_row.Input, ["string", "."])
        self.assertEqual(parsed_row.Output, None)
        self.assertEqual(parsed_row.Transform, {"Input": "mistake"})


class EvaluatingRowsMismatchTestCase2(unittest.TestCase):
    def setUp(self):
        self.raw_row = RawSheetEntry(
            "Resolver", "Fs", r'["string", "."]', "null", r'{"Input": "cwd"}'
        )

        self.transform = ("mistake", lambda x: [x[0], {"cwd": x[1]}])

    def test_parsing(self):
        parsed_row = self.raw_row.parse()

        parsed_row.transform(*self.transform)

        self.assertEqual(parsed_row.PartType, "Resolver")
        self.assertEqual(parsed_row.PartName, "Fs")
        self.assertListEqual(parsed_row.Input, ["string", "."])
        self.assertEqual(parsed_row.Output, None)
        self.assertEqual(parsed_row.Transform, {"Input": "cwd"})
